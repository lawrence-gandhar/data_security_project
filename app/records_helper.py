# Django authentication
from django.contrib.auth.models import User
from app.models import *
from django.core.paginator import Paginator
import openpyxl, os, sys, re
from django.conf import settings
from django.db import IntegrityError
from django.utils import timezone
import app.combiners_helper as combiners_helper

from django.db.models import Q, Count, Sum, Prefetch

#===========================================================================================
# Excel to DB insertion 
# ==========================================================================================
#

def insert_into_db(file_ins, file_path):
    path = settings.MEDIA_ROOT+file_path

    wb_obj = openpyxl.load_workbook(path, data_only = True) 
    sheet_obj = wb_obj.active 

    total_rows = sheet_obj.max_row 
    total_columns = sheet_obj.max_column

    msg = "Total Rows: {0}, Total Columns: {1}".format(total_rows- 1 , total_columns)

    # Loop will print all columns name 
    for row in range(2, total_rows + 1): 

        category = sheet_obj.cell(row = row, column = 1).value
        sub_category = sheet_obj.cell(row = row, column = 2).value
        brand = sheet_obj.cell(row = row, column = 3).value
        pe = sheet_obj.cell(row = row, column = 7).value

        cat, sub_cat, brand_ins, pe_ins = combiners_helper.category_sub_brand_insertion(category, sub_category, brand, pe)
       
        record = RecordsManagement(
            record_file = file_ins,
            is_active = file_ins.is_active,
            category = cat,
            sub_category = sub_cat,
            brand = brand_ins,
            previous_exhibition = pe_ins,
            contact_person = sheet_obj.cell(row = row, column = 4).value,
            contact_number = sheet_obj.cell(row = row, column = 5).value,
            email = sheet_obj.cell(row = row, column = 6).value,
        )

        record.save()
        
    return msg

#===========================================================================================
# RECORDS LIST
# ==========================================================================================
# 
def RecordsList(page = None, records_per_page = None, file_ins = None, kwargs = None):

    if file_ins is None:
        try:
            file_ins = FileSubmission.objects.exclude(is_active = False).latest('id')
        except:
            return False, False
    else:
        try:
            file_ins = FileSubmission.objects.get(pk = file_ins)
        except:
            return False, False

    records = RecordsManagement.objects.filter(record_file = file_ins)
    
    
    if kwargs is not None:
    
        if kwargs["active"] is not None:
            records = records.filter(is_active = kwargs["active"])
                
        if kwargs["assigned"] is not None:
            records = records.filter(is_assigned = kwargs["assigned"])
            
        if kwargs["cate"] is not None:
            records = records.filter(category_id = kwargs["cate"])
            
        if kwargs["sub_cat"] is not None:
            records = records.filter(sub_category_id = kwargs["sub_cat"])
            
        if kwargs["brand"] is not None:
            records = records.filter(brand_id = kwargs["brand"])

        if kwargs["search"] is not None:
            records = records.filter(brand__brand_name__icontains = kwargs["search"])
    
    
    records = records.select_related('category', 'sub_category', 'brand', 'record_file', 'previous_exhibition')
    records = records.values('id' ,'category__category_name', 'sub_category__category_name', 'brand__brand_name', 
                'contact_person', 'contact_number', 'email', 'is_active', 'record_file__uploaded_on', 'remarks', 'remark_added_on',
                'record_file__record_file_name', 'is_assigned', 'assigned_to', 'assigned_to__first_name', 'assigned_to__last_name' , 
                'assigned_on', 'is_completed','disposition', 'previous_exhibition__name').order_by('id')
    
    per_page = 10
    if records_per_page is not None:
        per_page = records_per_page

    if page is None:
        page = 1
    
    paginator = Paginator(records, per_page)
    records = paginator.get_page(page)

    return file_ins, records


#===========================================================================================
# RECORDS FILES LIST
# ==========================================================================================
#
def RecordsFileList():
    return FileSubmission.objects.all().values().order_by('-id')

#===========================================================================================
# Records Status
# ==========================================================================================
#
def RecordsStatus(records = list(), opt = '0'):
    for rec in records:
        try:
            record = RecordsManagement.objects.get(pk = int(rec))
            record.is_active = int(opt)
            record.save()
        except:
            pass


#===========================================================================================
# AUTO ASSIGNMENT Records  - All Staff
# ==========================================================================================
#
def AutoAssignRecords(file_ins = 0, opt = False, staff_list = [0]):
    if opt:
    
        if staff_list[0] == 0:
            app_perm = User.objects.filter(is_staff = True, is_superuser = False, is_active = True,)
        else:
            app_perm = User.objects.filter(pk__in = staff_list)
            
        app_perm = app_perm.select_related('app_permissions')        
        
                        
        try:
            file_ins = FileSubmission.objects.get(pk = int(file_ins))        
        except:
            file_ins = None
            return False
        
        
        for user in app_perm:
                    
            ret, buff_size = PreviousAssignmentsExists(user.id, file_ins)
            
            if not ret:
                
                records = RecordsManagement.objects.filter(is_active = True, record_file = file_ins, is_assigned = False,)
                 
                if len(user.app_permissions.dedicated_to_category.all()) > 0:
                    records = records.filter(category__in = user.app_permissions.dedicated_to_category.all())
                    
                
                if len(user.app_permissions.dedicated_to_sub_category.all()) > 0:
                    records = records.filter(sub_category__in = user.app_permissions.dedicated_to_sub_category.all())
                    
                if len(user.app_permissions.dedicated_to_brand.all()) > 0:
                    records = records.filter(brand__in = user.app_permissions.dedicated_to_brand.all())
                    
                if len(user.app_permissions.dedicated_to_pe.all()) > 0:
                    records = records.filter(previous_exhibition__in = user.app_permissions.dedicated_to_pe.all())
                                        
                if user.app_permissions.full_access:
                    records = records.values()
                else:
                    records = records.values()[:buff_size]
                    
                print(records.query)    
                
                for record in records:
                    try:
                        rec = RecordsManagement.objects.get(pk = record["id"])
                        rec.assigned_to_id = user.id
                        rec.assigned_on = timezone.now()
                        rec.is_assigned = opt
                        rec.save()  
                    except:
                        pass

#===========================================================================================
# CHECK PREVIOUS ASSIGNMENTS OF RECORDS TO USER IN THE SAME FILE/SLOT 
# ==========================================================================================
#
def PreviousAssignmentsExists(user_id, file_ins):

    app_perm = User.objects.get(pk = user_id)

    records = RecordsManagement.objects.filter(is_active = True, record_file = file_ins, is_assigned = True, 
                assigned_to = user_id) 
    
    completed_records = records.filter(is_completed = True).count()
    
    if completed_records >= app_perm.app_permissions.record_access_size:
        all_records = 0
    else:
        all_records = records.count()
    
    if all_records == app_perm.app_permissions.record_access_size and completed_records < app_perm.app_permissions.record_access_size:
        return True, 0
        
    return False, app_perm.app_permissions.record_access_size - all_records
        

#===========================================================================================
# CHECK PREVIOUS ASSIGNMENTS OF RECORDS TO USER IN THE SAME FILE/SLOT 
# ==========================================================================================
#
def GetRecord(user_id = None, page = None, records_per_page = None,  kwargs = None):
    records = RecordsManagement.objects.filter(assigned_to = user_id, is_completed = False)
    
    record_fetch = records.select_related('category', 'sub_category','brand', 'record_file','previous_exhibition')
    record_fetch = record_fetch.values('id', 'is_active' ,'category__category_name', 'sub_category__category_name', 'brand__brand_name', 
                'contact_person', 'contact_number', 'email', 'record_file__record_file_name','assigned_on', 'remarks', 
                'remark_added_on', 'disposition', 'previous_exhibition__name', 'is_completed').order_by('id')   

    per_page = 5
    if records_per_page is not None:
        per_page = records_per_page

    if page is None:
        page = 1
    
    paginator = Paginator(record_fetch, per_page)
    record_fetch = paginator.get_page(page)

    return record_fetch
    

#===========================================================================================
# RECORDS STAFF ASSIGN
# ==========================================================================================
#
def RecordsStaffAssign(records = list(), opt = '0', staff = '0'):
    for rec in records:
        try:
            if staff != '0' and opt != '0':            
                record = RecordsManagement.objects.get(pk = int(rec))
                record.is_assigned = True
                record.assigned_to_id = int(staff)
                record.assigned_on = timezone.now()
                record.save()
        except:
            pass


#===========================================================================================
# RECORDS STAFF ASSIGN
# ==========================================================================================
#
def RecordsApproval(file_ins = None, records = list(), opt = '0', staff = '0'):    

    if file_ins is not None:
    
        try:
            file_ins = FileSubmission.objects.get(pk = file_ins)
        except:
            return False
            
        if staff == '0':
            records = RecordsManagement.objects.filter(record_file = file_ins, is_completed = False, is_assigned = True, disposition__gt = 0).values('id')
        elif staff == '1':
            records = RecordsManagement.objects.filter(pk__in = records, is_completed = False, is_assigned = True, disposition__gt = 0 ).values('id')
        else:
            if len(records) > 0:
                records = RecordsManagement.objects.filter(record_file = file_ins, assigned_to_id = staff, is_completed = False, pk__in = records, disposition__gt =0 ).values('id')
            else:    
                records = RecordsManagement.objects.filter(record_file = file_ins, assigned_to_id = staff, is_completed = False, disposition__gt = 0).values('id')
        
        for rec in records:
            obj = RecordsManagement.objects.get(pk = rec["id"])
            obj.is_completed = True
            obj.save()
    
    
    
    
    
    
    
    
    
        